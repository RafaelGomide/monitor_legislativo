"""
Gera xlsx formatados com estilo profissional.
Paleta: preto/branco/cinza + azul/roxo para cabeçalhos e destaques.
"""

import tempfile
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

# ── Paleta ────────────────────────────────────────────────────────────────────
COR_HEADER_BG   = "1E1B4B"   # roxo escuro
COR_HEADER_FG   = "FFFFFF"
COR_LINHA_PAR   = "F4F4F8"   # cinza muito claro
COR_LINHA_IMPAR = "FFFFFF"
COR_BORDA       = "CCCCDD"

# Cores por resultado
COR_ALTERADO  = "FFDEDE"   # vermelho claro
COR_SEM_ALTER = "DEFFED"   # verde claro
COR_PRIMEIRA  = "FFF8DE"   # amarelo claro
COR_ERRO      = "FFE8D0"   # laranja claro

MAPA_RESULTADO = {
    "🔴 COM ALTERAÇÃO":      COR_ALTERADO,
    "✅ SEM ALTERAÇÃO":      COR_SEM_ALTER,
    "🆕 PRIMEIRA VERIFICAÇÃO": COR_PRIMEIRA,
    "❌ ERRO":               COR_ERRO,
}

FONT_NOME = "Arial"


def _borda():
    s = Side(style="thin", color=COR_BORDA)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color)


def salvar_resultado_xlsx(df: pd.DataFrame, nome_base: str = "resultado") -> str:
    """
    Gera um xlsx estilizado com a tabela de resultados.
    Retorna o caminho do arquivo temporário criado.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"

    cols = list(df.columns)
    n_cols = len(cols)

    # ── Título mesclado ───────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    titulo = ws["A1"]
    titulo.value = "MONITORAMENTO DE PROJETOS DE LEI — RESULTADO DA VERIFICAÇÃO"
    titulo.font      = Font(name=FONT_NOME, bold=True, size=12, color=COR_HEADER_FG)
    titulo.fill      = _fill(COR_HEADER_BG)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # ── Cabeçalho ─────────────────────────────────────────────────────────────
    header_font  = Font(name=FONT_NOME, bold=True, color=COR_HEADER_FG, size=9)
    header_fill  = _fill("2D2A6E")   # roxo médio
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[2].height = 32
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=2, column=ci, value=col)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = _borda()

    # ── Dados ─────────────────────────────────────────────────────────────────
    col_resultado = cols.index("RESULTADO") + 1 if "RESULTADO" in cols else None

    body_font = Font(name=FONT_NOME, size=9)

    for ri, row in df.iterrows():
        excel_row = ri + 3   # linha 1=título, 2=header, dados a partir de 3
        ws.row_dimensions[excel_row].height = 38

        resultado_val = str(row.get("RESULTADO", ""))
        row_color = MAPA_RESULTADO.get(resultado_val,
                    COR_LINHA_PAR if excel_row % 2 == 0 else COR_LINHA_IMPAR)

        for ci, col in enumerate(cols, 1):
            valor = row[col]
            if pd.isna(valor):
                valor = ""

            cell = ws.cell(row=excel_row, column=ci, value=str(valor))
            cell.font   = body_font
            cell.border = _borda()

            # Célula RESULTADO: cor específica + negrito
            if ci == col_resultado:
                cell.fill = _fill(row_color)
                cell.font = Font(name=FONT_NOME, size=9, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
            else:
                # Fundo levemente colorido nas linhas com alteração
                if resultado_val in ("🔴 COM ALTERAÇÃO", "❌ ERRO"):
                    cell.fill = _fill("FFF5F5")
                else:
                    bg = COR_LINHA_PAR if excel_row % 2 == 0 else COR_LINHA_IMPAR
                    cell.fill = _fill(bg)

                if col in ("EMENTA", "DESCRIÇÃO", "OBSERVAÇÃO"):
                    cell.alignment = Alignment(horizontal="left", vertical="top",
                                               wrap_text=True)
                elif col in ("SETOR", "CASA", "Nº PL", "ÚLTIMA TRAMITAÇÃO", "ÓRGÃO"):
                    cell.alignment = Alignment(horizontal="center", vertical="center",
                                               wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center",
                                               wrap_text=True)

    # ── Aba de resumo ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumo")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 12

    resumo_header = Font(name=FONT_NOME, bold=True, color=COR_HEADER_FG, size=10)
    resumo_fill   = _fill(COR_HEADER_BG)

    ws2.merge_cells("A1:B1")
    t = ws2["A1"]
    t.value     = "RESUMO DA VERIFICAÇÃO"
    t.font      = Font(name=FONT_NOME, bold=True, color=COR_HEADER_FG, size=11)
    t.fill      = resumo_fill
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    for ci, h in enumerate(["Resultado", "Quantidade"], 1):
        cell = ws2.cell(row=2, column=ci, value=h)
        cell.font      = resumo_header
        cell.fill      = _fill("2D2A6E")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _borda()
    ws2.row_dimensions[2].height = 24

    contagem = df["RESULTADO"].value_counts() if "RESULTADO" in df.columns else {}
    for ri2, (label, count) in enumerate(contagem.items(), 3):
        cor = MAPA_RESULTADO.get(label, COR_LINHA_IMPAR)
        for ci2, val in enumerate([label, count], 1):
            cell = ws2.cell(row=ri2, column=ci2, value=val)
            cell.font      = Font(name=FONT_NOME, size=10,
                                  bold=(ci2 == 1))
            cell.fill      = _fill(cor)
            cell.alignment = Alignment(horizontal="center" if ci2 == 2 else "left",
                                       vertical="center")
            cell.border    = _borda()
        ws2.row_dimensions[ri2].height = 22

    # Total
    total_row = len(contagem) + 3
    ws2.cell(total_row, 1, "TOTAL").font  = Font(name=FONT_NOME, bold=True, size=10)
    ws2.cell(total_row, 1).fill          = _fill("E8E8F0")
    ws2.cell(total_row, 1).border        = _borda()
    ws2.cell(total_row, 1).alignment     = Alignment(horizontal="left", vertical="center")
    ws2.cell(total_row, 2, len(df)).font = Font(name=FONT_NOME, bold=True, size=10)
    ws2.cell(total_row, 2).fill          = _fill("E8E8F0")
    ws2.cell(total_row, 2).border        = _borda()
    ws2.cell(total_row, 2).alignment     = Alignment(horizontal="center", vertical="center")

    # ── Larguras (aba Resultado) ───────────────────────────────────────────────
    larguras = {
        "SETOR":              20,
        "CASA":               10,
        "Nº PL":              24,
        "EMENTA":             45,
        "ÚLTIMA TRAMITAÇÃO":  18,
        "ÓRGÃO":              12,
        "SITUAÇÃO ATUAL":     35,
        "DESCRIÇÃO":          45,
        "RESULTADO":          24,
        "OBSERVAÇÃO":         35,
    }
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = larguras.get(col, 18)

    ws.freeze_panes  = "A3"
    ws2.freeze_panes = "A3"

    # ── Salva ─────────────────────────────────────────────────────────────────
    tmp = tempfile.NamedTemporaryFile(
        suffix=f"_{nome_base}.xlsx", delete=False, prefix="monitor_"
    )
    tmp.close()
    wb.save(tmp.name)
    return tmp.name


def salvar_planilha_xlsx(df: pd.DataFrame, nome_base: str = "planilha_atualizada") -> str:
    """
    Salva a planilha atualizada (com STATUS MAIS RECENTE) como xlsx formatado.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "2026 Monitoramento - PL"

    cols = list(df.columns)

    # Título
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    t = ws["A1"]
    t.value     = "MONITORAMENTO PARLAMENTAR DE PROJETOS DE LEI"
    t.font      = Font(name=FONT_NOME, bold=True, size=12, color=COR_HEADER_FG)
    t.fill      = _fill(COR_HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Cabeçalho
    h_font  = Font(name=FONT_NOME, bold=True, color=COR_HEADER_FG, size=9)
    h_fill  = _fill("2D2A6E")
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 36

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=2, column=ci, value=col)
        cell.font = h_font; cell.fill = h_fill
        cell.alignment = h_align; cell.border = _borda()

    # Dados
    b_font = Font(name=FONT_NOME, size=9)
    for ri, row in df.iterrows():
        excel_row = ri + 3
        ws.row_dimensions[excel_row].height = 50
        bg = COR_LINHA_PAR if excel_row % 2 == 0 else COR_LINHA_IMPAR

        for ci, col in enumerate(cols, 1):
            val = row[col]
            if pd.isna(val): val = ""
            cell = ws.cell(row=excel_row, column=ci, value=str(val))
            cell.font   = b_font
            cell.fill   = _fill(bg)
            cell.border = _borda()
            cell.alignment = Alignment(
                horizontal="center" if col in ("CÂMARA /SENADO", "N° - PL", "APRESENTAÇÃO") else "left",
                vertical="top", wrap_text=True
            )

    # Larguras
    larg_planilha = {
        "SETOR": 22, "CÂMARA /SENADO": 14, "N° - PL": 22,
        "AUTOR": 26, "APRESENTAÇÃO": 16, "EMENTA": 55,
        "FORMA DE APRECIAÇÃO": 30, "REGIME DE TRAMITAÇÃO": 30,
        "LINK PARA ACOMPANHAMENTO": 55, "PROCESSO SEI": 22,
        "REPRESENTAÇÃO  SETORIAL": 35, "STATUS MAIS RECENTE": 55,
    }
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = larg_planilha.get(col, 20)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}2"

    tmp = tempfile.NamedTemporaryFile(
        suffix=f"_{nome_base}.xlsx", delete=False, prefix="monitor_"
    )
    tmp.close()
    wb.save(tmp.name)
    return tmp.name
